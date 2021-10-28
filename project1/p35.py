import random
from openpyxl import load_workbook
workbook = load_workbook(filename="game.xlsx")
workbook.sheetnames
sheet = workbook.active
sheet.title
for j in range(5):
    sheet["A"+str(random.randrange(1,21))]=0
    sheet["B"+str(random.randrange(1,21))]=0
    sheet["C"+str(random.randrange(1,21))]=0
    sheet["D"+str(random.randrange(1,21))]=0
    sheet["E"+str(random.randrange(1,21))]=0
    sheet["F"+str(random.randrange(1,21))]=0
    sheet["G"+str(random.randrange(1,21))]=0
    sheet["H"+str(random.randrange(1,21))]=0
    sheet["I"+str(random.randrange(1,21))]=0
    sheet["J"+str(random.randrange(1,21))]=0
    sheet["K"+str(random.randrange(1,21))]=0
    sheet["L"+str(random.randrange(1,21))]=0
workbook.save(filename="game.xlsx")
c=0
def play():
    global c
    z=input("จงเดาตัวเลขมาส่ะ : ")
    if sheet[z].value==0:
        print("GAME OVER!!!")
        print("Your Score : "+str(c))
        sheet.delete_cols(1,21)
        workbook.save(filename="game.xlsx")
    elif sheet[z].value!=0:
        c=c+1
        play()
play()
    

    