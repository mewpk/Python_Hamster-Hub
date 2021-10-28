
import random
from openpyxl import load_workbook
workbook = load_workbook(filename="game.xlsx")
workbook.sheetnames
sheet = workbook.active
sheet.title
def mew():
    for j in range(5):
        sheet["A"+str(random.randrange(1,11))]=0
        sheet["B"+str(random.randrange(1,11))]=0
        sheet["C"+str(random.randrange(1,11))]=0
        sheet["D"+str(random.randrange(1,11))]=0
        sheet["E"+str(random.randrange(1,11))]=0
        workbook.save(filename="game.xlsx")
    play()
c=0
s=0
py=[]
def play():
    global c
    global s
    global u
    for p in range(int(u)):
        py.append("Player "+str(p+1))
    z=input("จงเดาตัวเลขมาส่ะ : ")
    if sheet[z].value==0:
        print("GAME OVER!!!")
        print(str(py[s])+"  Your Score : "+str(c))
        s=s+1
        sheet.delete_cols(1,21)
        workbook.save(filename="game.xlsx")
        if s < int(u):
            mew()
    elif sheet[z].value!=0:
        c=c+1
        play()
u=input("Player : ")
mew()