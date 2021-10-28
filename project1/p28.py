from openpyxl import load_workbook

workbook = load_workbook(filename="helo.xlsx")

workbook.sheetnames
sheet = workbook.active
sheet.title

for x in range(1,11):
    sheet["A"+str(x)] = x
    workbook.save(filename="helo.xlsx")
    print(sheet["A"+str(x)].value)