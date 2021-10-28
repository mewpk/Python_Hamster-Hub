from openpyxl import load_workbook

workbook = load_workbook(filename="helo.xlsx")

workbook.sheetnames
sheet = workbook.active
sheet.title

x = sheet["A1"].value
print(x)

sheet["A2"] = "hello"
sheet["B2"] = "world!"

workbook.save(filename="helo.xlsx")

print(sheet["A2"].value,sheet["B2"].value)