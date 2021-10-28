from openpyxl import load_workbook
workbook = load_workbook(filename="score.xlsx")
workbook.sheetnames
sheet = workbook.active
sheet.title
mew=[]
kanan=[]
pan=0
aipan=0
for x in range(5,25):
    del mew[:]
    mew.append(sheet["B"+str(x)].value)
    mew.append(sheet["C"+str(x)].value)
    mew.append(sheet["D"+str(x)].value) 
    mew.append(sheet["E"+str(x)].value) 
    mew.append(sheet["F"+str(x)].value) 
    mew.append(sheet["G"+str(x)].value)
    sheet["I"+str(x)] =sum(mew)
    workbook.save(filename="score.xlsx")
    if sum(mew)>80:
        sheet["J"+str(x)] =  "A"
        workbook.save(filename="score.xlsx")
        pan=pan+1
    elif sum(mew)>70:
        sheet["J"+str(x)] =  "B"
        workbook.save(filename="score.xlsx")
        pan=pan+1
    elif sum(mew)>60:
        sheet["J"+str(x)] =  "C"
        workbook.save(filename="score.xlsx")
        pan=pan+1
    elif sum(mew)>=50:
        sheet["J"+str(x)] =  "D"
        workbook.save(filename="score.xlsx")
        pan=pan+1
    elif sum(mew)<50:
        sheet["J"+str(x)] =  "F"
        workbook.save(filename="score.xlsx")
        aipan=aipan+1
for y in range(5,25):
    kanan.append(sheet["I"+str(y)].value)
    sheet["M"+str(11)] =  min(kanan)
    sheet["M"+str(12)] =  max(kanan)
    sheet["M"+str(13)] =  sum(kanan)//19
    sheet["M"+str(14)] =  str(pan)
    sheet["M"+str(15)] =  str(aipan)
    workbook.save(filename="score.xlsx")
print("เสร็จ")