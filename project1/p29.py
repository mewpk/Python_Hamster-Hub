from openpyxl import load_workbook

workbook = load_workbook(filename="helo.xlsx")
workbook.sheetnames
sheet = workbook.active
sheet.title
z= 0
c=0
v=0
for x in range(1,101):
    if x <=25:
        sheet["A"+str(x)] = x
        workbook.save(filename="helo.xlsx")
        print(sheet["A"+str(x)].value)
    elif x<=50:
        z = z+1
        sheet["B"+str(z)] = x
        workbook.save(filename="helo.xlsx")
        print(sheet["B"+str(z)].value)
    elif x<=75:
        c= c+1
        sheet["C"+str(c)] = x
        workbook.save(filename="helo.xlsx")
        print(sheet["C"+str(c)].value)
            
    elif x<=100:
        v=v+1
        sheet["D"+str(v)] = x
        workbook.save(filename="helo.xlsx")
        print(sheet["D"+str(v)].value)
