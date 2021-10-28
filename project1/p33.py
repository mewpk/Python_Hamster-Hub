from openpyxl import load_workbook

workbook = load_workbook(filename="score1.xlsx")

workbook.sheetnames
sheet = workbook.active
sheet.title
for i in range(5,25):
    average = 0
    for row in sheet.iter_rows(min_row=i, min_col=2, max_row=i, max_col=7):
        for cell in row:
            if (cell.internal_value != None):
                average += int(cell.internal_value)
                sheet["I"+str(i)] = average
    if average >= 80 and average <= 100:
        sheet["J" + str(i)] = "A"
    elif average >= 70 and average <= 80:
        sheet["J" + str(i)] = "B"
    elif average >= 60 and average <= 70:
        sheet["J" + str(i)] = "C"
    elif average >= 50 and average <= 60:
        sheet["J" + str(i)] = "D"
    elif average <= 50:
        sheet["J" + str(i)] = "F"

workbook.save(filename = 'score.xlsx')

for h in range(0,21):
    x = sheet["J" + str(i)].value
    print(x)
for j in range(0,21):
    y = sheet["I" + str(i)].value
    print(y)